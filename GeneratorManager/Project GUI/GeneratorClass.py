import pandas as pd
import numpy as np
from datetime import datetime
from os import path
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

# generator list
genList = ['1A', '2A', '3A', '4A', '1B', '2B', '3B', '4B']

# categories to be saved in the XLSX file
data_dictionary = {'StartDate': [], 'StartTime': [], 'StopDate': [], 'StopTime': [], 'RunHourStart': [], \
                   'RunHourStop': [], 'Load': [], 'ReasonForRun': [], 'RunTime': [], 'CalcFuel': []}
# These variables are going to be used in 'generator' function
capacity, derate, quater_load, half_load, threequat_load, full_load = 0, 0, 0, 0, 0, 0


# function for creating XLSX file in the storage
def mkexcel(genName, dataframe):
    if not path.isfile('Gen_Data.xlsx'):
        wb = openpyxl.Workbook()
        wb1 = wb['Sheet']
        wb1.title = genName
        for r in dataframe_to_rows(dataframe, index=False, header=True):
            wb1.append(r)

        return wb.save(filename='Gen_Data.xlsx')
    else:
        wb = openpyxl.load_workbook(filename='Gen_Data.xlsx')
        try:
            df = pd.read_excel('Gen_Data.xlsx', sheet_name=genName)
            dataframe = df.append(dataframe, ignore_index=False)
            wb1 = wb.get_sheet_by_name(genName)
            wb.remove_sheet(wb1)
            wb1 = wb.create_sheet(genName)
        except:
            wb1 = wb.create_sheet(genName)
        for r in dataframe_to_rows(dataframe, index=False, header=True):
            wb1.append(r)
        return wb.save(filename='Gen_Data.xlsx')


# function for bringing the information of generators
def generator(genName):
    global capacity, derate, quater_load, half_load, threequat_load, full_load, cur_runHour
    if genName in ['1A', '2A', '3A', '1B', '2B', '3B']:
        capacity, derate, quater_load, half_load, threequat_load, full_load = 2000, 1795, 43.00, 71.00, 103.00, 135.00
    else:
        capacity, derate, quater_load, half_load, threequat_load, full_load = 2000, 1550, 46.50, 82.00, 107.30, 141.30
    # get the cumulative run hour from the last entered data
    try:
        temp_df = pd.read_excel('Gen_Data.xlsx', sheet_name='GEN ' + genName)
        cur_runHour = temp_df.iloc[len(temp_df) - 1]['RunHourStop']
    except:
        # print('\nThere is no data for the cumulative run hours.\nPlease enter the cumulative run hours.')
        # cur_runHour = input()
        # while not cur_runHour.isdigit():
        #     print('\nPlease enter correct number in hour.')
        #     cur_runHour = input()
        # cur_runHour = make a field for cumulative run hours or a separate diaglog popped up when user enters the submit
        cur_runHour = 3
        cur_runHour = float(cur_runHour)


# function to convert the input date into datetime format, and to save as list form
def inputdate(date_entry):
    try:
        date = datetime.strptime(date_entry, "%m-%d-%y")
        return [date.strftime("%x")]
    except:
        print(r"[Error] Please use valid form in mm-dd-yy")
        inputdate(input())


# function to convert the input time into datetime format, and to save as list form
def inputtime(time_entry):
    try:
        date = datetime.strptime(time_entry, "%H:%M")
        return [date.strftime("%X")[:5]]
    except:
        print(r"[Error] Please use valid form in hh:mm")
        inputtime(input())


# function to calculate the difference between 'start time' and 'stop time'
def runtime(sdate, stime, edate, etime):
    tempdate1 = sdate + " " + stime
    tempdate2 = edate + " " + etime
    print('Date1:',tempdate1)
    print('Date2:',tempdate2)
    diff = datetime.strptime(tempdate2, "%m/%d/%y %H:%M") - datetime.strptime(tempdate1, "%m/%d/%y %H:%M")
    days = diff.days
    totalSeconds = diff.seconds
    hours, remainder = divmod(totalSeconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    hours += days * 24
    return ["{:02d}:{:02d}".format(hours, minutes)]


# function to convert %H:%M format into hours
def cumrunhour(runHour_now, runtime):
    i, j = map(int, runtime[0].split(':'))
    return [round(runHour_now + i + j / 60, 1)]


# function to calculate the amount of fuel consumed based on 'run time'
def calcfuel(value, runtimee):
    if value / derate < 0.25:
        i, j = map(int, runtimee.split(":"))
        return [round((i * quater_load) + (j * quater_load / 60), 2)]
    elif value / derate < 0.5:
        i, j = map(int, runtimee.split(":"))
        return [round((i * half_load) + (j * half_load / 60), 2)]
    elif value / derate < 0.75:
        i, j = map(int, runtimee.split(":"))
        return [round((i * threequat_load) + (j * threequat_load / 60), 2)]
    else:
        i, j = map(int, runtimee.split(":"))
        return [round((i * full_load) + (j * full_load / 60), 2)]

def checkCapacity(generatorname):
    genName = generatorname.upper()
    generator(genName)
    wholestr = "Capacity: {}\nDerate: {}\n1/4 Load: {}\n1/2 Load: {}\n3/4 Load: {}\nFull Load: {}" \
        .format(capacity, derate, quater_load, half_load, threequat_load, full_load)
    return wholestr


def dumpData(genname, startdate, starttime, enddate, endtime, loadvalue, reasonRun):
    cur_runHour = 0
    data_dictionary['StartDate'] = startdate
    data_dictionary['StartTime'] = starttime
    data_dictionary['StopDate'] = enddate
    data_dictionary['StopTime'] = endtime

    print('SDate',startdate)
    print('STime', starttime)
    print('Edate', enddate)
    print('Etime', endtime)

    # sdate, stime, edate, etime = data_dictionary['StartDate'][0], data_dictionary['StartTime'][0], \
    #                              data_dictionary['StopDate'][0], data_dictionary['StopTime'][0]

    # data_dictionary['RunTime'] = runtime(sdate, stime, edate, etime)
    data_dictionary['RunTime'] = runtime(startdate, starttime, enddate, endtime)

    # runhourstart
    data_dictionary['RunHourStart'] = [cur_runHour]

    # runhourstop
    cur_runHour = cumrunhour(cur_runHour, data_dictionary['RunTime'])
    data_dictionary['RunHourStop'] = cur_runHour

    data_dictionary['Load'] = [int(loadvalue)]

    data_dictionary['ReasonForRun'] = [reasonRun]

    # calculate fuel
    value = int(data_dictionary['Load'][0])
    runtimee = data_dictionary['RunTime'][0]
    data_dictionary['CalcFuel'] = calcfuel(value, runtimee)

    # create excel file
    genName = 'GEN ' + genname
    df = pd.DataFrame(data_dictionary)
    mkexcel(genName, df)
    secondresult = 'The excel file has been created.'
    return secondresult


# function to filter by date and to sum of the sorted values
def calsumhour(genName, month, year):
    ans = '00:00'
    # filter by year
    tempDF = df_dict[genName][pd.DatetimeIndex(df_dict[genName]['StartDate']).year == year]
    # fliter by month
    tempDF = tempDF[pd.DatetimeIndex(tempDF['StartDate']).month == month]
    tempDF = tempDF.reset_index(drop=True)
    if len(tempDF.index) == 1:
        return tempDF['RunTime'][0]
    elif len(tempDF.index) > 1:
        temp = tempDF['RunTime'].str.split(':')
        hourVal, minuteVal = 0, 0
        for i in temp:
            hourVal += int(i[0])
            minuteVal += int(i[1])
        ans = "{:02d}:{:02d}".format(hourVal, minuteVal)
        return ans
    else:
        return ans

# function to calculate the total hours
# def totalhour(genName):
#     totHour = 0
#     totMinute = 0
#     for i in df_MonthlyHour[genName]:
#         totHour += int(i.split(':')[0])
#         totMinute += int(i.split(':')[1])
#     totHour += totMinute / 60
#     return totHour

df_dict = {}

def showData(year):
    year = int(year)
    # import all generator data
    genList = ['GEN 1A', 'GEN 2A', 'GEN 3A', 'GEN 4A', 'GEN 1B', 'GEN 2B', 'GEN 3B', 'GEN 4B']
    for gen in genList:
        try:
            df_dict[gen] = pd.read_excel('Gen_Data.xlsx', sheet_name=gen, parse_dates=['StartDate', 'StopDate'])
        except:
            pass

        # create dictionary for monthly run hours
        monthlyRunHour = {}

        # calculate the sum of run hours per month
        for gen in genList:
            monthlySum = []
            try:
                len(df_dict[gen])
                for mon in range(1, 13):
                    try:
                        monthlySum.append(calsumhour(gen, mon, year))
                    except:
                        monthlySum.append('01:00')
            except:
                monthlySum = ['00:00', '00:00', '00:00', '00:00', '00:00', '00:00', '00:00', '00:00', '00:00', '00:00', '00:00',
                              '00:00']
            monthlyRunHour[gen] = monthlySum

        # create data frame for monthly run hours
        df_MonthlyHour = pd.DataFrame(monthlyRunHour,
                                      index=['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August',
                                             'September', 'October', 'November', 'December'])
        # create dictionary for totals
        totals = {}

        # calculate the total values
        for gen in genList:
            totalList = []
            # total run hours
            # totalList.append(totalhour(gen))
            totalList.append(0)
            # total fuel usage

            totalList.append(0)
            # total NOx
            if gen in ['GEN 4A', 'GEN 4B']:
                totalList.append(totalList[0] * 34.11)
            else:
                totalList.append(totalList[0] * 35.43)
            # total SO2
            if gen in ['GEN 4A', 'GEN 4B']:
                totalList.append(totalList[0] * 0.71)
            else:
                totalList.append(totalList[0] * 0.97)
            # total Co
            if gen in ['GEN 4A', 'GEN 4B']:
                totalList.append(totalList[0] * 1.16)
            else:
                totalList.append(totalList[0] * 2.25)
            # total PM 10
            if gen in ['GEN 4A', 'GEN 4B']:
                totalList.append(totalList[0] * 0.26)
            else:
                totalList.append(totalList[0] * 0.64)
            # total VOC
            if gen in ['GEN 4A', 'GEN 4B']:
                totalList.append(totalList[0] * 0.2)
            else:
                totalList.append(totalList[0] * 0.4)

            # add to dictionary
            totals[gen] = totalList

        # create data frame for the total values
        df_Totals = pd.DataFrame(totals, index=['Total Run Hours', 'Total Fuel Usage', 'Total NOx', 'Total SO2', 'Total CO',
                                                'Total PM 10', 'Total VOC']).round(2)

        # join two data frames
        df_summary = pd.concat([df_MonthlyHour, df_Totals])

        # export data frame as excel
        wb = openpyxl.Workbook()
        wb1 = wb['Sheet']
        wb1.title = 'Summary{}'.format(year)
        for r in dataframe_to_rows(df_summary, index=True, header=True):
            if r == [None]:
                pass
            else:
                wb1.append(r)
        wb.save(filename="Summary_{}.xlsx".format(year))

        print("\nThe 'Summary_{}.xlsx' file has been created.".format(year))
        sumaryfile = "The Summary.xlsx file has been created"
    return sumaryfile


